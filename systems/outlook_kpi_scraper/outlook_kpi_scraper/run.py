"""
Outlook KPI Scraper – main entry point.

Pipeline:
  1) Load env / config / keywords
  2) Fetch Outlook messages (COM)
  3) Score & filter candidates (deny lists, attachment boosts, penalties)
  4) For each candidate:
     a) Download attachments → parse for KPIs (attachment-first)
     b) Fall back to body-text regex extraction
     c) Route entity via entity_aliases
  5) Enforce data-integrity rules (skip empty-KPI rows)
  6) Batch-append qualifying rows to Google Sheets (with 429 backoff)
  7) Write RUN LOG PACK to logs/runs/<run_id>/
"""

import argparse
import logging
import os
import time
import traceback

from outlook_kpi_scraper.config import (
    load_all_keywords,
    load_sender_allowlist,
    load_entity_aliases,
    validate_startup_config,
)
from outlook_kpi_scraper.outlook_reader import OutlookReader
from outlook_kpi_scraper.filters import filter_candidates
from outlook_kpi_scraper.entity_router import route_entity
from outlook_kpi_scraper.kpi_extractor import extract_kpis, has_kpi_values, compute_confidence
from outlook_kpi_scraper.attachment_extractor import extract_kpis_from_attachments, get_attachment_decisions
from outlook_kpi_scraper.llm_extractor import (
    llm_available, extract_kpis_with_llm, merge_llm_into_regex,
)
from outlook_kpi_scraper.dep_check import check_ocr_dependencies
from outlook_kpi_scraper.ledger import Ledger
from outlook_kpi_scraper.run_logger import RunLogger
from outlook_kpi_scraper.source_matcher import (
    load_source_mapping, match_email, validate_extracted_kpis,
)
from outlook_kpi_scraper.writers.google_sheets_writer import GoogleSheetsWriter
from outlook_kpi_scraper.writers.csv_writer import CSVWriter
from outlook_kpi_scraper.utils import load_env
from outlook_kpi_scraper.quarantine_triage import triage_quarantined_emails, triage_available
from outlook_kpi_scraper.attachment_gate import evaluate_attachment_gate

log = logging.getLogger(__name__)

# ------------------------------------------------------------------
# Body-text financial signal detector
# ------------------------------------------------------------------

# Regex patterns that indicate the body contains actual financial data
# worth sending to the LLM. At least _BODY_SIGNAL_THRESHOLD hits required.
import re as _re

_DOLLAR_RE = _re.compile(r"\$[\d,]+(?:\.\d{2})?")           # $1,234.56
_PERCENT_RE = _re.compile(r"\d+\.?\d*\s*%")                 # 92.3%
_KPI_KEYWORD_RE = _re.compile(
    r"\b(?:revenue|cash\s*balance|bank\s*balance|pipeline|"
    r"closings?|occupancy|census|net\s*income|operating\s*income|"
    r"total\s*revenue|gross\s*revenue|net\s*revenue|"
    r"ending\s*balance|current\s*balance|"
    r"ebitda|noi|cap\s*rate|"
    r"total\s*expenses|total\s*income|"
    r"rent\s*roll|monthly\s*report|financial\s*summary)\b",
    _re.IGNORECASE,
)
_BODY_SIGNAL_THRESHOLD = 2  # need at least 2 distinct signal types

# Subject patterns that indicate deal discussion / legal negotiation
# rather than operating KPI reports.  Body-text LLM is skipped for these.
_DEAL_DISCUSSION_RE = _re.compile(
    r"\b(?:PSA|purchase\s+(?:and\s+)?sale|term\s*sheets?|data\s*room|"
    r"due\s*diligence|LOI|letter\s+of\s+intent|earnest\s+money|"
    r"closing\s+(?:documents?|instructions?)|escrow|"
    r"diligence\s+items?|title\s+(?:update|commitment|policy|search)|"
    r"wire\s+(?:instructions?|approval))\b",
    _re.IGNORECASE,
)


def _body_has_kpi_signals(body: str) -> bool:
    """Fast heuristic: does *body* contain enough financial signals
    to justify an LLM call?  Returns True if ≥ 2 signal types found
    among {dollar amounts, percentages, KPI keywords}."""
    signals = 0
    if _DOLLAR_RE.search(body):
        signals += 1
    if _PERCENT_RE.search(body):
        signals += 1
    if _KPI_KEYWORD_RE.search(body):
        signals += 1
    return signals >= _BODY_SIGNAL_THRESHOLD


def _is_deal_discussion(msg: dict) -> bool:
    """Return True if the email subject indicates deal discussion,
    legal negotiation, or transactional activity rather than an
    operating KPI report."""
    subject = (msg.get("subject") or "")
    return bool(_DEAL_DISCUSSION_RE.search(subject))


# ------------------------------------------------------------------
# Debug attachment mode (standalone, no Outlook needed)
# ------------------------------------------------------------------

def _debug_attachment(file_path: str):
    """Analyse a single attachment file and print suitability + KPI extraction results."""
    import json as _json

    # Set up minimal console logging
    logging.basicConfig(level=logging.DEBUG,
                        format="%(levelname)s %(name)s %(message)s")

    print(f"\n{'='*60}")
    print(f"  DEBUG ATTACHMENT: {file_path}")
    print(f"{'='*60}")

    if not os.path.exists(file_path):
        print(f"ERROR: File not found: {file_path}")
        return

    ext = os.path.splitext(file_path)[1].lower()
    filename = os.path.basename(file_path)
    file_size = os.path.getsize(file_path)
    print(f"  File: {filename}")
    print(f"  Size: {file_size:,} bytes")
    print(f"  Extension: {ext}")

    # ---- OCR deps ----
    from outlook_kpi_scraper.dep_check import check_ocr_dependencies
    ocr_deps = check_ocr_dependencies()
    print(f"\n  OCR deps: tesseract={ocr_deps['tesseract']}  poppler={ocr_deps['poppler']}  opencv={ocr_deps['opencv']}")

    # ---- Extract text ----
    text = ""
    used_ocr = False
    sheetnames: list[str] = []

    if ext == ".pdf":
        from outlook_kpi_scraper.ocr_service import extract_pdf_text_with_fallback
        text, used_ocr = extract_pdf_text_with_fallback(file_path)
    elif ext in (".xlsx",):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheetnames = wb.sheetnames
            parts = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    parts.append(" ".join(str(c) for c in row if c is not None))
            text = "\n".join(parts)
            wb.close()
        except Exception as exc:
            print(f"  XLSX parse error: {exc}")
    elif ext == ".xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(file_path)
            sheetnames = wb.sheet_names()
            parts = []
            for ws in wb.sheets():
                for row_num in range(ws.nrows):
                    parts.append(" ".join(
                        str(ws.cell_value(row_num, col))
                        for col in range(ws.ncols)
                        if ws.cell_value(row_num, col)
                    ))
            text = "\n".join(parts)
        except Exception as exc:
            print(f"  XLS parse error: {exc}")
    elif ext == ".csv":
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                text = f.read()
        except Exception as exc:
            print(f"  CSV read error: {exc}")
    else:
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                text = f.read(50000)
        except Exception:
            print(f"  Cannot read file as text")

    print(f"\n  Extracted text length: {len(text)} chars")
    print(f"  Used OCR: {used_ocr}")
    if sheetnames:
        print(f"  Sheet names: {sheetnames}")

    # ---- Suitability ----
    from outlook_kpi_scraper.kpi_suitability import compute_suitability
    suit = compute_suitability(
        text, filename=filename, sheetnames=sheetnames,
        is_pdf=(ext == ".pdf"),
        text_is_empty=(len(text.strip()) < 200),
    )
    print(f"\n  Suitability score: {suit['score']}")
    print(f"  Suitability tier:  {suit['tier']}")
    print(f"  Accept: {suit['accept_bool']}")
    print(f"  OCR candidate: {suit['used_ocr_candidate_bool']}")
    print(f"  Reasons:")
    for r in suit["reasons"]:
        print(f"    - {r}")
    if suit["reject_hits"]:
        print(f"  Reject hits: {suit['reject_hits']}")

    # ---- KPI extraction (uses the same logic as attachment_extractor._scan_row) ----
    import re as _re
    from outlook_kpi_scraper.kpi_labels import match_label, _REVERSE
    from outlook_kpi_scraper.kpi_extractor import parse_money, parse_percent

    _MONEY_DBG = _re.compile(r"[\$]?\s*[\-\(]?\s*[\d,]+\.?\d*\s*[kKmMbB]?\s*\)?")

    def _dbg_parse_value(raw, field):
        if not raw:
            return None
        raw = raw.strip()
        if field == "occupancy":
            if "%" in raw:
                return parse_percent(raw)
            v = parse_money(raw)
            if v is not None and 0 < v <= 100:
                return v / 100.0
            return v
        if "count" in field:
            v = parse_money(raw)
            return int(v) if v is not None else None
        return parse_money(raw)

    kpi: dict = {}
    kpi_evidence: list[str] = []

    for line_num, line in enumerate(text.splitlines(), 1):
        parts = _re.split(r"[:\t|]+|\s{2,}", line)
        for i, cell in enumerate(parts):
            cell_stripped = cell.strip()
            if not cell_stripped:
                continue
            field = match_label(cell_stripped)
            if field is None:
                sub = _re.split(r"[:\-=]\s*", cell_stripped, maxsplit=1)
                if len(sub) == 2:
                    field = match_label(sub[0])
                    if field and field not in kpi:
                        val = _dbg_parse_value(sub[1], field)
                        if val is not None:
                            kpi[field] = val
                            ev_line = line.strip()[:120]
                            kpi_evidence.append(f"line{line_num}: {field}={val} evidence='{ev_line}'")
                continue
            if field in kpi:
                continue
            found = False
            for j in range(i + 1, min(i + 4, len(parts))):
                val = _dbg_parse_value(parts[j].strip(), field)
                if val is not None:
                    kpi[field] = val
                    ev_line = line.strip()[:120]
                    kpi_evidence.append(f"line{line_num}: {field}={val} evidence='{ev_line}'")
                    found = True
                    break
            # Fallback: label and value share same cell (PDF financial layouts)
            if not found:
                cell_lower = cell_stripped.lower()
                for syn in sorted((s for s, f in _REVERSE.items() if f == field), key=len, reverse=True):
                    idx = cell_lower.find(syn)
                    if idx >= 0:
                        after = cell_stripped[idx + len(syn):]
                        m = _MONEY_DBG.search(after)
                        if m:
                            val = _dbg_parse_value(m.group(0).strip(), field)
                            if val is not None:
                                kpi[field] = val
                                ev_line = line.strip()[:120]
                                kpi_evidence.append(f"line{line_num}: {field}={val} (same-cell) evidence='{ev_line}'")
                        break

    print(f"\n  Extracted KPIs:")
    if kpi:
        for k, v in kpi.items():
            print(f"    {k}: {v}")
    else:
        print(f"    (none)")

    print(f"\n  KPI Evidence:")
    if kpi_evidence:
        for ev in kpi_evidence:
            print(f"    {ev}")
    else:
        print(f"    (no evidence)")

    # First 500 chars of text
    print(f"\n  Text preview (first 500 chars):")
    print(f"  {text[:500]}")
    print(f"\n{'='*60}")


def _write_llm_candidates(run_dir, extracted_rows, quarantined):
    """Write llm_candidates.csv listing docs for LLM enrichment and quarantined emails for triage.

    Two sections:
      category=tier1_extraction / tier2_extraction — matched emails with Tier 1/2
          attachments whose KPIs should be enriched by GPT-4o.
      category=quarantine_triage — unmatched emails that need lightweight
          LLM classification.
    """
    import csv
    import re as _re
    out_path = os.path.join(run_dir, "llm_candidates.csv")
    rows = []

    # 1) Tier 1/2 from extracted rows (parse tier from evidence)
    for _eid, kpi_row in extracted_rows:
        ev = kpi_row.get("evidence_source", "") or ""
        tier_m = _re.search(r"tier=(\d+)", ev)
        tier = int(tier_m.group(1)) if tier_m else 4
        if tier <= 2:
            rows.append({
                "category": f"tier{tier}_extraction",
                "sender": kpi_row.get("sender", ""),
                "subject": kpi_row.get("subject", ""),
                "entity": kpi_row.get("entity", ""),
                "tier": tier,
                "source_rule": kpi_row.get("source_rule_id", ""),
                "kpis_found": ";".join(
                    f for f in ("revenue", "cash", "pipeline_value",
                                "closings_count", "orders_count", "occupancy")
                    if kpi_row.get(f) is not None
                ),
                "llm_ran": "LLM:" in ev,
                "evidence_snippet": ev[:200],
            })

    # 2) Quarantined emails for triage
    for msg in quarantined:
        rows.append({
            "category": "quarantine_triage",
            "sender": msg.get("sender_email", ""),
            "subject": msg.get("subject", ""),
            "entity": "",
            "tier": "",
            "source_rule": "",
            "kpis_found": "",
            "llm_ran": False,
            "evidence_snippet": f"score={msg.get('candidate_score', 0)} "
                                f"reasons={msg.get('candidate_reason', [])}",
        })

    if not rows:
        return

    fieldnames = ["category", "sender", "subject", "entity", "tier",
                  "source_rule", "kpis_found", "llm_ran", "evidence_snippet"]
    try:
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        log.info("Wrote %d LLM candidates to %s (tier1/2: %d, quarantine: %d)",
                 len(rows), out_path,
                 sum(1 for r in rows if r["category"] != "quarantine_triage"),
                 sum(1 for r in rows if r["category"] == "quarantine_triage"))
    except Exception as exc:
        log.warning("Failed to write llm_candidates.csv: %s", exc)


def main():
    parser = argparse.ArgumentParser(
        description="Scrape Outlook KPI emails and append to Google Sheet."
    )
    parser.add_argument("--days", type=int, default=7, help="Days to look back")
    parser.add_argument("--mailbox", type=str, required=True, help="Mailbox display name")
    parser.add_argument("--folder", type=str, default="Inbox", help="Folder name (single)")
    parser.add_argument("--folders", type=str, default=None,
                        help="Comma-separated folder names (e.g. 'Inbox,Sent Items,Junk Email')")
    parser.add_argument("--max", type=int, default=200, help="Max messages to process (per folder)")
    parser.add_argument("--subfolder-days", type=int, default=None,
                        help="Days to look back for subfolders (e.g. Inbox/PAYROLL). "
                             "Defaults to --days if not set. Use a longer window "
                             "(e.g. 365) for infrequently-updated subfolders.")
    parser.add_argument("--debug", action="store_true", help="Enable debug output")
    parser.add_argument("--require-kpi", action="store_true", default=True,
                        help="Only append rows with at least one KPI value (default: True)")
    parser.add_argument("--no-require-kpi", dest="require_kpi", action="store_false",
                        help="Append all extracted rows even if no KPI values")
    parser.add_argument("--batch-size", type=int, default=200,
                        help="Google Sheets batch size (rows per API call)")
    parser.add_argument("--debug-attachment", type=str, default=None, metavar="PATH",
                        help="Debug a single attachment file: print suitability, OCR status, and extracted KPIs then exit.")
    args = parser.parse_args()

    # ---- Debug attachment mode (runs standalone, no Outlook needed) ----
    if args.debug_attachment:
        _debug_attachment(args.debug_attachment)
        return

    debug = args.debug or os.environ.get("DEBUG", "0") in ("1", "true", "True")
    t0 = time.time()

    # ---- Resolve folder list ----
    if args.folders:
        folder_list = [f.strip() for f in args.folders.split(",") if f.strip()]
    else:
        folder_list = [args.folder]

    # ---- Environment & logging ----
    env = load_env()
    run_logger = RunLogger()   # creates logs/runs/<run_id>/ and sets up file logging
    log.info("=== Outlook KPI Scraper – run %s ===", run_logger.run_id)
    log.info("Args: days=%d mailbox=%s folders=%s max=%d debug=%s require_kpi=%s",
             args.days, args.mailbox, folder_list, args.max, debug, args.require_kpi)

    # ---- Startup validation (config + dependencies) ----
    validate_startup_config()

    # ---- OCR dependency check (non-fatal) ----
    ocr_status = check_ocr_dependencies()
    log.info("OCR deps: tesseract=%s poppler=%s opencv=%s",
             ocr_status.get('tesseract'), ocr_status.get('poppler'), ocr_status.get('opencv'))

    # ---- Config ----
    keywords = load_all_keywords()
    sender_allowlist = load_sender_allowlist()
    entity_aliases = load_entity_aliases()
    log.info("Loaded %d keywords, %d allowlist entries, %d entity alias groups",
             len(keywords), len(sender_allowlist),
             len(entity_aliases.get("keywords", {})) + len(entity_aliases.get("sender_domains", {})))

    # ---- Fetch messages ----
    ledger = Ledger()
    reader = OutlookReader(
        mailbox=args.mailbox, folder=folder_list,
        days=args.days, max_items=args.max,
        subfolder_days=args.subfolder_days,
    )
    try:
        messages = reader.fetch_messages()
    except Exception as exc:
        log.error("FATAL: fetch_messages crashed: %s", exc)
        messages = []
    scanned = len(messages)
    log.info("Fetched %d messages from Outlook", scanned)

    # ---- Filter candidates ----
    candidates = []
    non_candidates = []
    for msg in messages:
        # Track domain for tuning suggestions
        sender_email = (msg.get("sender_email") or "")
        if "@" in sender_email:
            run_logger.track_domain(sender_email.split("@")[-1].lower())

        if ledger.is_processed(msg["entry_id"]):
            continue
        is_candidate = filter_candidates(
            msg, keywords, sender_allowlist,
            debug=debug,
            has_attachments=msg.get("has_attachments", False),
            has_kpi_attachment=msg.get("has_kpi_attachment", False),
        )
        if is_candidate:
            candidates.append(msg)
            run_logger.add_candidate(
                msg,
                score=msg.get("candidate_score", 0),
                reasons=msg.get("candidate_reason", []),
                has_attachments=msg.get("has_attachments", False),
                attachment_names=msg.get("attachment_names", ""),
            )
        else:
            non_candidates.append(msg)

    log.info("Candidates: %d / %d scanned", len(candidates), scanned)

    # ---- Source Mapping ----
    source_cfg = load_source_mapping()
    source_rules_count = len(source_cfg.get("sources", []))
    log.info("Source mapping: %d rules loaded", source_rules_count)

    quarantined = []

    # ---- Thread deduplication ----
    # Group candidates by ConversationTopic so we only extract KPIs
    # from the MOST RECENT email in each thread (avoids duplicate
    # rows from quoted reply chains).
    _seen_conversations: dict[str, str] = {}   # conv_topic → entry_id of newest
    _thread_dupes: set[str] = set()             # entry_ids to skip

    for msg in candidates:
        conv_topic = (msg.get("conversation_topic") or "").strip()
        eid = msg["entry_id"]
        if conv_topic:
            if conv_topic in _seen_conversations:
                # We already have a newer message for this thread
                _thread_dupes.add(eid)
            else:
                _seen_conversations[conv_topic] = eid

    skipped_thread_dedup = len(_thread_dupes)
    if skipped_thread_dedup:
        log.info("Thread dedup: %d duplicate thread emails will be skipped",
                 skipped_thread_dedup)

    # ---- Extract KPIs ----
    extracted_rows = []      # list of (entry_id, kpi_row)
    failed_extractions = []
    skipped_no_kpi = 0
    skipped_quarantine = 0
    skipped_kpi_validation = 0
    skipped_noise = 0        # attachment gate noise rejections
    skipped_deal_discussion = 0  # deal-discussion subject filter

    for msg in candidates:
        entry_id = msg["entry_id"]
        try:
            # 0-pre) Thread dedup — skip older replies in same conversation
            if entry_id in _thread_dupes:
                log.debug("Thread dedup skip: sender=%s subject=%s",
                          msg.get("sender_email"), msg.get("subject"))
                continue

            # 0a) Attachment type gate — deterministic pre-filter
            #     Prevents image-only noise from consuming extraction time
            att_gate = evaluate_attachment_gate(msg)
            msg["attachment_gate"] = att_gate["decision"]

            if att_gate["decision"] in ("NOISE_IMAGE_ONLY", "NOISE_SIGNATURE", "NOISE_SUBJECT"):
                skipped_noise += 1
                run_logger.add_quarantined(
                    msg,
                    reason=f"attachment_gate:{att_gate['decision']} – {att_gate['reason']}",
                    top_scores=[],
                )
                log.info("NOISE SKIP: sender=%s subject=%s gate=%s",
                         msg.get("sender_email"), msg.get("subject"),
                         att_gate["decision"])
                quarantined.append(msg)
                continue

            # 0b) Source mapping — deterministic rule matching
            src_match = match_email(msg)

            if not src_match.matched:
                # Unknown source → quarantine (do NOT blind parse)
                quarantined.append(msg)
                skipped_quarantine += 1
                run_logger.add_quarantined(
                    msg,
                    reason="no source rule matched",
                    top_scores=src_match.all_scores[:3],
                )
                log.info("QUARANTINE: sender=%s subject=%s (no source rule)",
                         msg.get("sender_email"), msg.get("subject"))
                continue

            # Use entity from source rule if available, else fall back to alias router
            entity = src_match.entity or route_entity(msg, entity_aliases)

            # 1) Attachment-first extraction
            att_kpis = None
            source_type = "body"
            raw_item = reader.get_raw_item(entry_id)
            if raw_item and msg.get("has_kpi_attachment"):
                try:
                    att_kpis = extract_kpis_from_attachments(
                        raw_item, entry_id, run_logger.attachments_dir,
                    )
                    if att_kpis:
                        source_type = "attachment"
                        log.info("Attachment KPIs for entry_id=%s: %s",
                                 entry_id[-12:], {k: v for k, v in att_kpis.items() if k != "evidence"})
                except Exception as exc:
                    log.warning("Attachment extraction failed for %s: %s", entry_id[-12:], exc)
                    run_logger.add_extraction_failure(
                        sender=msg.get("sender_email", ""),
                        subject=msg.get("subject", ""),
                        error=str(exc),
                    )

            # 1b) Body-text LLM extraction — runs when attachments yielded
            #     no KPIs AND the body text contains financial signals
            #     (dollar amounts, percentages, or KPI keywords).
            #     Skipped entirely for deal-discussion subjects (PSA, Term
            #     Sheet, Data Room, etc.) to prevent third-party acquisition
            #     figures from being misclassified as operating KPIs.
            body_llm_kpis = None
            body_text = (msg.get("body") or "").strip()
            _att_empty = att_kpis is None or not any(
                att_kpis.get(f) is not None
                for f in ("revenue", "cash", "pipeline_value",
                          "closings_count", "orders_count", "occupancy")
            )

            _is_deal = _is_deal_discussion(msg)
            if _is_deal:
                skipped_deal_discussion += 1
                log.debug("Deal discussion skip (body LLM): sender=%s subject=%s",
                          msg.get("sender_email"), msg.get("subject"))

            _body_has_financial_signal = False
            if _att_empty and not _is_deal and len(body_text) > 80:
                _body_has_financial_signal = _body_has_kpi_signals(body_text)

            if _att_empty and _body_has_financial_signal and llm_available():
                try:
                    body_llm_kpis = extract_kpis_with_llm(
                        body_text,
                        doc_type="email_body",
                        filename=f"body:{msg.get('sender_email', '')} / {msg.get('subject', '')}",
                    )
                    if body_llm_kpis:
                        source_type = "body_llm"
                        log.info("Body-text LLM KPIs for entry_id=%s: %s",
                                 entry_id[-12:],
                                 {f: body_llm_kpis[f].get("value")
                                  for f in body_llm_kpis
                                  if body_llm_kpis[f].get("value") is not None})
                except Exception as exc:
                    log.warning("Body-text LLM extraction failed for %s: %s",
                                entry_id[-12:], exc)

            # 2) Body-text regex extraction (fills gaps)
            kpi_row = extract_kpis(msg, entity, attachment_kpis=att_kpis)

            # 2b) Merge body LLM results into kpi_row (overrides regex gaps)
            if body_llm_kpis:
                evidence_list = kpi_row.get("evidence_source", "").split("; ")
                kpi_row = merge_llm_into_regex(
                    kpi_row, body_llm_kpis, evidence_list, source="body",
                )
                kpi_row["evidence_source"] = "; ".join(evidence_list)

            # 3) Data integrity gate
            if args.require_kpi and not has_kpi_values(kpi_row):
                skipped_no_kpi += 1
                run_logger.add_skipped_candidate(
                    msg,
                    score=msg.get("candidate_score", 0),
                    reasons=msg.get("candidate_reason", []),
                    why_skipped="no KPI values",
                )
                log.debug("Skipped (no KPI values): sender=%s subject=%s",
                          msg.get("sender_email"), msg.get("subject"))
                continue

            # 3b) Per-source KPI validation
            kpi_validation = validate_extracted_kpis(kpi_row, src_match)
            if not kpi_validation["valid"]:
                skipped_kpi_validation += 1
                missing = kpi_validation["missing_required"]
                run_logger.add_skipped_candidate(
                    msg,
                    score=msg.get("candidate_score", 0),
                    reasons=msg.get("candidate_reason", []),
                    why_skipped=f"required KPIs missing: {', '.join(missing)} (rule={src_match.rule_id})",
                )
                log.warning("KPI validation REJECT: rule=%s missing=%s sender=%s",
                            src_match.rule_id, missing, msg.get("sender_email"))
                continue

            # 4) Compute confidence and attach metadata
            confidence = compute_confidence(kpi_row)
            att_name = att_kpis.get("attachment_names", "") if att_kpis else ""
            proof = kpi_row.get("evidence_source", "")[:200]

            # Populate sheet-output metadata fields
            kpi_row["run_id"] = run_logger.run_id
            kpi_row["message_id"] = entry_id[-24:] if entry_id else ""
            kpi_row["sender"] = msg.get("sender_email", "")
            kpi_row["subject"] = msg.get("subject", "")
            kpi_row["candidate_score"] = msg.get("candidate_score", 0)
            kpi_row["candidate_reasons"] = ";".join(msg.get("candidate_reason", []))
            kpi_row["source_type"] = source_type
            kpi_row["attachment_name"] = att_name
            if source_type == "attachment" and att_name:
                kpi_row["evidence_snippet"] = f"attachment: {att_name}"
            elif kpi_row.get("evidence_source"):
                kpi_row["evidence_snippet"] = kpi_row["evidence_source"][:120]
            else:
                kpi_row["evidence_snippet"] = ""
            kpi_row["extractor_version"] = "v3.0-source-mapped"
            kpi_row["confidence"] = confidence
            # Validation flags: anomaly alerts if any
            kpi_row["validation_flags"] = kpi_row.get("alerts", "")

            # Source mapping metadata
            kpi_row["source_rule_id"] = src_match.rule_id
            kpi_row["source_match_score"] = round(src_match.match_score, 3)
            kpi_row["source_report_type"] = src_match.report_type
            kpi_row["source_parse_confidence"] = kpi_validation.get("parse_confidence", 0.0)

            extracted_rows.append((entry_id, kpi_row))
            run_logger.add_extracted_row(
                kpi_row,
                sender_email=msg.get("sender_email", ""),
                subject=msg.get("subject", ""),
                evidence_source=kpi_row.get("evidence_source", ""),
                source_type=source_type,
                attachment_name=att_name,
                extraction_proof=proof,
                confidence_score=confidence,
                entry_id=entry_id,
                source_rule_id=src_match.rule_id,
                source_match_score=src_match.match_score,
            )

            if debug:
                log.info("Extracted: sender=%s entity=%s revenue=%s cash=%s pipeline=%s source=%s rule=%s",
                         msg.get("sender_email"), entity,
                         kpi_row.get("revenue"), kpi_row.get("cash"),
                         kpi_row.get("pipeline_value"),
                         kpi_row.get("evidence_source", ""),
                         src_match.rule_id)

        except Exception as exc:
            tb = traceback.format_exc()
            failed_extractions.append({
                "sender": msg.get("sender_email") or msg.get("sender_name"),
                "subject": msg.get("subject"),
                "received": msg.get("received_dt"),
                "error": str(exc),
                "traceback": tb,
            })
            run_logger.add_extraction_failure(
                sender=msg.get("sender_email", ""),
                subject=msg.get("subject", ""),
                error=str(exc),
            )
            log.error("Extraction error: sender=%s subject=%s error=%s",
                      msg.get("sender_email"), msg.get("subject"), exc)

    log.info("Extracted: %d rows, skipped_no_kpi=%d, quarantined=%d, "
             "noise_skipped=%d, kpi_validation_rejects=%d, extraction_errors=%d, "
             "thread_dedup=%d, deal_discussion_skipped=%d",
             len(extracted_rows), skipped_no_kpi, skipped_quarantine,
             skipped_noise, skipped_kpi_validation, len(failed_extractions),
             skipped_thread_dedup, skipped_deal_discussion)

    # ---- Write LLM candidates manifest ----
    # Captures Tier 1/2 docs (for enrichment) + quarantined emails (for triage)
    _write_llm_candidates(run_logger.run_dir, extracted_rows, quarantined)

    # ---- Quarantine triage (GPT-4o-mini classification) ----
    triage_summary = {}
    if quarantined and triage_available():
        log.info("Starting quarantine triage on %d emails...", len(quarantined))
        triage_summary = triage_quarantined_emails(
            quarantined, run_logger.run_dir,
        )
        log.info("Triage summary: %s", triage_summary)
    elif quarantined:
        log.info("Quarantine triage skipped (LLM not available). "
                 "%d emails written to quarantined.csv only.", len(quarantined))

    # ---- Write to Google Sheets (batched) ----
    writer = None
    if env.get("GOOGLE_SHEET_ID") and env.get("GOOGLE_SERVICE_ACCOUNT_JSON_PATH"):
        writer = GoogleSheetsWriter(env, batch_size=args.batch_size)
    else:
        writer = CSVWriter()

    for entry_id, row in extracted_rows:
        if ledger.is_processed(entry_id):
            continue
        writer.append_row(row)

    # Flush batch (Google Sheets) or no-op (CSV already wrote per-row)
    appended = 0
    failed_appends_count = 0
    if hasattr(writer, "flush"):
        appended, failed_appends_count = writer.flush()
        # Record per-row results from the writer
        for r in writer.results:
            run_logger.add_append_result(**r)
    else:
        appended = len(extracted_rows)

    # Mark processed in ledger
    for entry_id, row in extracted_rows:
        ledger.mark_processed(entry_id, msg=row)

    # ---- Summary ----
    failed_count = len(failed_extractions) + failed_appends_count
    duration = time.time() - t0

    run_logger.set_summary(
        scanned=scanned,
        candidate_count=len(candidates),
        extracted_count=len(extracted_rows),
        appended_count=appended,
        failed_count=failed_count,
        skipped_no_kpi=skipped_no_kpi,
        quarantined_count=skipped_quarantine,
        noise_skipped=skipped_noise,
        kpi_validation_rejects=skipped_kpi_validation,
        duration_sec=duration,
        args={
            "days": args.days,
            "mailbox": args.mailbox,
            "folders": folder_list,
            "max": args.max,
            "debug": debug,
            "require_kpi": args.require_kpi,
            "batch_size": args.batch_size,
            "source_rules": source_rules_count,
            "triage_summary": triage_summary,
        },
    )

    # Collect attachment decisions from the extractor module
    att_decisions = get_attachment_decisions()
    run_logger.flush(attachment_decisions=att_decisions)

    # Console summary
    chip_review_path = os.path.abspath(os.path.join(run_logger.run_dir, "CHIP_REVIEW.txt"))
    print(f"\n{'='*60}")
    print(f"  RUN COMPLETE: {run_logger.run_id}")
    print(f"{'='*60}")
    print(f"  scanned={scanned}  candidates={len(candidates)}  "
          f"extracted={len(extracted_rows)}  appended={appended}  "
          f"failed={failed_count}  skipped_no_kpi={skipped_no_kpi}  "
          f"noise_skipped={skipped_noise}  "
          f"quarantined={skipped_quarantine}  kpi_rejects={skipped_kpi_validation}  "
          f"thread_dedup={skipped_thread_dedup}  deal_skipped={skipped_deal_discussion}  "
          f"duration={duration:.1f}s")
    if triage_summary and triage_summary.get("classified", 0) > 0:
        print(f"\n  TRIAGE: classified={triage_summary['classified']}  "
              f"financial_hits={triage_summary.get('financial_count', 0)}  "
              f"by_label={triage_summary.get('by_label', {})}  "
              f"cost=~${triage_summary.get('cost_estimate_usd', 0):.4f}")
    print(f"\n  CHIP REVIEW:  {chip_review_path}")
    print(f"  Run log pack: {os.path.abspath(run_logger.run_dir)}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
